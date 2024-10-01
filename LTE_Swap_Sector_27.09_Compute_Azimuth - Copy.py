import os
import re
import time
import math
import shutil
import pymssql
import openpyxl
import warnings
import datetime
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

"""Define Input & Output folder path."""
input_path              = r"C:\BOT\LTE_Swap Sector_Sep24\LTE_Swap_Sector_20Sep24"
output_path             = r"C:\BOT\LTE_Swap Sector_Sep24\LTE_Swap_Sector_20Sep24"

"""Define Input & Output file path & name."""
# Define file paths
input_file              = os.path.join(input_path, "itk_ho_stats.xlsx")
input_user              = os.path.join(input_path, "Input_User.xlsx")
source_output           = os.path.join(input_path, "Source_CI.xlsx")
template_file           = os.path.join(input_path, "Input_Swap_Sector_Template_4G.xlsx")
pre_post_output         = os.path.join(input_path, "Pre_Post.xlsx")
siteid_file             = os.path.join(input_path, "SiteID.xlsx")
pivot_output            = os.path.join(input_path, "Pivot.xlsx")
direction_file          = os.path.join(input_path, "Direction.xlsx")
lte_ho_analysis_file    = os.path.join(input_path, 'LTE_HO_Analysis.xlsx')
pm_data_file            = os.path.join(input_path, "itk_pm_stats.xlsx")


"""error handling."""
def handle_file_not_found(file_path):
    print(f"Error: File not found: {file_path}")

"""copy siteid from 'Input_Swap_Sector_Template_4G' will use while saving output file."""  
def get_site_id():

    # Path to the Excel file
    workbook_path = os.path.join(input_path, 'Input_User.xlsx')
    
    try:
        # Load the workbook and select the sheet
        workbook = load_workbook(workbook_path)
        if 'input' not in workbook.sheetnames:
            raise ValueError("The worksheet 'input' does not exist in the workbook.")
        
        sheet = workbook['input']
        
        # Read the value from cell B5
        site_id = sheet['B5'].value
        
        if site_id is None:
            raise ValueError("Cell B5 is empty or does not contain a valid value.")
        
        return site_id
    
    except Exception as e:
        print(f"An error occurred: {e}")
        return None
    
"""Extract source information from a cell value."""
def calculate_source(cell):
    try:
        equal_index = cell.find('=')
        left_part = cell[equal_index + 1:equal_index + 7]
        comma_index = cell.find(',')
        right_part = cell[comma_index - 12:comma_index - 9]
        return "{}_{}".format(left_part, right_part)
    except Exception as e:
        print(f"Error calculating source: {e}")
        return ""

"""Process the input data and save to source_output."""
def process_data():

    try:
        df = pd.read_excel(input_file, sheet_name='1', header=0)
        df['Source'] = df['object'].apply(calculate_source)
        df['Target'] = df['object'].apply(lambda x: x.split('=')[-1].strip())
        df.rename(columns={'time': 'Date', 'Attempts': 'Attempts'}, inplace=True)
        output_df = df[['Date', 'Source', 'Target', 'Attempts']]
        output_df.to_excel(source_output, index=False)
        print(f"Output file saved: {source_output}")
    except FileNotFoundError:
        handle_file_not_found(input_file)
    except Exception as e:
        print(f"An error occurred in process_data: {e}")
        
"""Define pre and post periods based on the template file and save to pre_post_output."""
def define_pre_post_periods():
    try:
        df_source = pd.read_excel(source_output)
        df_template = pd.read_excel(input_user)

        pre_end_date = pd.to_datetime(df_template.loc[df_template['Input'] == 'Pre_Period_End', 'Input Required'].values[0], format='%d-%m-%Y')
        post_start_date = pd.to_datetime(df_template.loc[df_template['Input'] == 'Post_Period_Start', 'Input Required'].values[0], format='%d-%m-%Y')

        df_source['Date'] = pd.to_datetime(df_source['Date'])
        df_source['Period'] = df_source['Date'].apply(lambda x: 'Pre' if x < pre_end_date else ('Post' if x >= post_start_date else ''))

        df_source.dropna(subset=['Period'], inplace=True)
        df_source.to_excel(pre_post_output, index=False)
        print(f"Pre-post period data saved successfully in '{pre_post_output}'.")
    except FileNotFoundError:
        handle_file_not_found(source_output)
    except Exception as e:
        print(f"An error occurred in define_pre_post_periods: {e}")        
        
"""Delete rows with empty periods and save the file."""
def delete_buffer_periods_from_input():

    try:
        df = pd.read_excel(pre_post_output)
        df = df.dropna(subset=['Period'])
        df['Date'] = pd.to_datetime(df['Date']).dt.strftime('%d-%m-%Y')
        df.to_excel(pre_post_output, index=False)  # Overwrite the input file
        print("Empty period rows are deleted and the file is saved.")
    except FileNotFoundError:
        handle_file_not_found(pre_post_output)
    except Exception as e:
        print(f"An error occurred in delete_empty_periods: {e}")
        

"""Fetch eNodeBID from SQL and delete apptempts = o in pre-post.xlsx."""
def Deltion_Zero_Attempts():
    
    # Load the data from the Excel file
    df = pd.read_excel(pre_post_output, sheet_name='Sheet1')

    # Step 2: Delete entries where Attempts = 0 or blank
    df = df[(df['Attempts'] != 0) & (df['Attempts'].notna())]

    # Save the modified DataFrame back to Excel
    df.to_excel(pre_post_output, index=False)

    print("Entries with Attempts = 0 or blank have been removed and the file has been saved.")


"""Create the Input_Swap_Sector_Template_4G.xlsx if it doesn't exist."""
def create_template_file():    
    if not os.path.exists(template_file):
        with pd.ExcelWriter(template_file, engine='openpyxl') as writer:
            empty_df = pd.DataFrame(columns=['GCID', 'CellName'])
            empty_df.to_excel(writer, sheet_name='NodebId', index=False)
            empty_df.to_excel(writer, sheet_name='Azimuth', index=False)
            print(f"Template file created at: {template_file}")
    else:
        print(f"Template file already exists at: {template_file}")

"""Load data from an Excel file."""
def load_excel_data(file_path, sheet_name=0):
    return pd.read_excel(file_path, sheet_name=sheet_name)


"""Reads database credentials from an input_user.xlsx file."""
def read_credentials_from_excel(filename: str) -> dict:

    df = pd.read_excel(filename, sheet_name="input")
    col_map = {'Server': 'server', 'User': 'user', 'Password': 'password'}
    
    credentials = {}
    for item in df.itertuples(index=False):
        key = item[0]  # Get the 'Input' column value
        if key in col_map:
            credentials[col_map[key]] = re.sub(r'[()\']', '', str(item[1]))

    return credentials


"""sql query to fetch the enodebid and cellname fron heidi sql database."""
def execute_query1(conn):
    query1 = """
    SELECT 
        CONCAT('2621', '-', Ericsson_BB.dbo.ENodebfunction.enbid, '-', Ericsson_BB.dbo.EUtranCellFDD.cellid) AS GCID, 
        Ericsson_BB.dbo.EUtranCellFDD.s_cell AS CellName 
    FROM 
        Ericsson_BB.dbo.ENodebfunction 
    INNER JOIN 
        Ericsson_BB.dbo.EUtranCellFDD ON Ericsson_BB.dbo.ENodebfunction.s_site = Ericsson_BB.dbo.EUtranCellFDD.s_site 
    WHERE 
        Ericsson_BB.dbo.EUtranCellFDD.tranid > CONVERT(VARCHAR(8), DATEADD(DAY, -30, GETDATE()), 112)
    GROUP BY 
        CONCAT('2621', '-', Ericsson_BB.dbo.ENodebfunction.enbid, '-', Ericsson_BB.dbo.EUtranCellFDD.cellid), 
        Ericsson_BB.dbo.EUtranCellFDD.s_cell;
    """
    return pd.read_sql(query1, conn)

"""Save DataFrame to an Excel file."""
def save_to_excel(dataframe, file_path, sheet_name):    
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
        if sheet_name in writer.book.sheetnames:
            # Remove the existing sheet
            std = writer.book[sheet_name]
            writer.book.remove(std)

        # Write the new data to the sheet
        dataframe.to_excel(writer, sheet_name=sheet_name, index=False, header=True)

        
"""Process the NodebId data in nput_Swap_Sector_Template_4G.xlsx and generate SiteID_NEID."""
def site_id_neid():
    sheet_name = 'NodebId'
    
    try:
        # Read the sheet with headers
        df = pd.read_excel(template_file, sheet_name=sheet_name)

        # Check if the second row contains duplicate headers and remove it
        if df.shape[0] > 1 and df.iloc[1, 0] == "GCID":
            df = df.drop(index=1).reset_index(drop=True)

        # Generate SiteID_NEID based on CellName
        def process_cell_name(cell_name):
            first_part = cell_name[:6]
            match = re.search(r'_(LA|LB|LG|LK|LU|LH|LT|LD)(\d+)', cell_name)
            second_part = match.group(0)[1:] if match else ''
            site_id_neid = f"{first_part}_{second_part}"

            if site_id_neid and site_id_neid[-1].isalpha() and site_id_neid[-1].islower():
                site_id_neid = site_id_neid[:-1]
            
            return site_id_neid

        df['SiteID_NEID'] = df['CellName'].apply(process_cell_name)
        final_df = df[['GCID', 'SiteID_NEID', 'CellName']]
        
        # Save the DataFrame back to the same sheet
        save_to_excel(final_df, template_file, sheet_name)

    except FileNotFoundError:
        print(f"Error: The file '{template_file}' was not found.")
    except Exception as e:
        print(f"An error occurred: {e}")

        
"""Prepare the list of tarket GCID from pre-post.xlsx and fetch the GCIS and cell name for them only."""
def eNodeBID_Fetch():
    # Step 1: Create template file if it doesn't exist
    create_template_file()
    
    # Step 2: Read data from Pre_Post.xlsx
    try:
        pre_post_data = load_excel_data(pre_post_output, sheet_name='Sheet1')
        raw_gcid = pre_post_data['Target'].drop_duplicates().tolist()
        
        # Step 3: Read SQL credentials and connect
        credentials = read_credentials_from_excel(input_user)
        conn = pymssql.connect(server=credentials['server'], user=credentials['user'], password=credentials['password'])

        # Suppress warnings
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            # Step 4: Execute SQL query1
            query1_df = execute_query1(conn)

        # Step 5: Filter query1 results based on RAWGCID
        filtered_df = query1_df[query1_df['GCID'].isin(raw_gcid)]

        # Step 6: Save filtered results to Excel
        save_to_excel(filtered_df, template_file, 'NodebId')

        # Step 7: Process the NodebId sheet
        site_id_neid()
        
    except FileNotFoundError as e:
        print(f"File not found: {e}")
    except Exception as e:
        print(f"An error occurred: {e}")
    finally:
        conn.close()  # Properly close the connection
        
        
"""Perform vlookup to match SiteID_NEID from Input_Swap_Sector_Template_4G.xlsx and save to siteid_file."""
def neighbour_gcid_siteid():

    try:
        pivot_df = pd.read_excel(pre_post_output)
        template_df = pd.read_excel(template_file, sheet_name='NodebId')

        merged_df = pivot_df.merge(template_df[['GCID', 'SiteID_NEID']], left_on='Target', right_on='GCID', how='left')
        merged_df.rename(columns={'Target': 'Target_GCID', 'SiteID_NEID': 'Target'}, inplace=True)
        merged_df.drop(columns=['GCID'], inplace=True)
        merged_df = merged_df[['Date', 'Source', 'Target', 'Attempts', 'Period', 'Target_GCID']]
        merged_df = merged_df[~merged_df['Target_GCID'].astype(str).str.startswith('2622')]
        merged_df.to_excel(siteid_file, index=False)
        print(f"Output saved successfully in '{siteid_file}'.")
    except FileNotFoundError:
        handle_file_not_found(pre_post_output)
    except Exception as e:
        print(f"An error occurred in vlookup_site_ne: {e}")
        
"""Identify missing GCID and save to Missing_GCID.xlsx."""
def missing_gcid_enode_database():

    try:
        df = pd.read_excel(siteid_file, sheet_name='Sheet1')
        filtered_df = df[df['Target'].isnull()]
        selected_df = filtered_df[['Target_GCID', 'Attempts']]
        pivot_table = pd.pivot_table(selected_df, values='Attempts', index='Target_GCID', aggfunc='sum')
        pivot_table.reset_index(inplace=True)
        pivot_table.to_excel(input_path + r'\Missing_GCID.xlsx', index=False)
        print("missing_gcid.xlsx saved..'")
    except FileNotFoundError:
        handle_file_not_found(siteid_file)
    except Exception as e:
        print(f"An error occurred in missing_gcid: {e}")
        
"""Prepare and save the pivot table to pivot_output."""
def prepare_ho_pre_post_relationwise():

    try:
        df = pd.read_excel(siteid_file)
        pivot_table = pd.pivot_table(df, index=['Source', 'Target'], columns='Period', values='Attempts', aggfunc='mean', fill_value=0)
        pivot_table.columns = ['Post_avg_Attempts', 'Pre_avg_Attempts']
        pivot_table.reset_index(inplace=True)
        pivot_table = pivot_table.sort_values(by='Pre_avg_Attempts', ascending=False)[['Source', 'Target', 'Pre_avg_Attempts', 'Post_avg_Attempts']]
        pivot_table.to_excel(pivot_output, index=False)
        print(f"prepare_ho_pre_post_relationwise excel saved successfully as '{pivot_output}'.")
    except FileNotFoundError:
        handle_file_not_found(siteid_file)
    except Exception as e:
        print(f"An error occurred in prepare_pivot_table: {e}")


"""Read input_user and established connection with sql."""
def read_credentials_from_excel_Sitedatabse(filename: str) -> dict:
    """Reads database credentials from an Excel file."""
    df = pd.read_excel(filename, sheet_name="input")
    col_map = {'Server': 'server', 'User': 'user', 'Password': 'password'}
    
    credentials = {}
    for item in df.itertuples(index=False):
        key = item[0]  # Get the 'Input' column value
        if key in col_map:
            credentials[col_map[key]] = re.sub(r'[()\']', '', str(item[1]))

    return credentials

def execute_query2(connection, query2):
    """Executes the given SQL query and returns the result as a DataFrame."""
    return pd.read_sql(query2, connection)


"""Fetch sitedatabase from heidi using query only for the Sites which are present in Source & Traget in Pivot.xlsx."""
def site_database():
    try:
        # Step 1: Read the Pivot Excel file
        df = pd.read_excel(pivot_output, sheet_name='Sheet1')

        # Step 2: Copy Source and Target to one list and remove duplicates
        source_target_list = pd.concat([df['Source'], df['Target']]).unique().tolist()

        # Step 3: Take the first 6 characters of each item in the list
        site_list = [site[:6] for site in source_target_list]

        # Step 4: Construct SQL query
        sql_query2 = """
        SELECT 
            CASE 
                WHEN LEFT(Band, 3) = 'NGM' THEN 'NR'
                ELSE LEFT(Band, 3)
            END AS Technology,
            REPLACE(INDENT_NO, ']', '') AS [Site ID],
            CAST(CELL_NODE_ID AS INT) AS [Cell ID],
            NE_ID AS [CI],  
            CAST(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(SECTOR_NO, 'A', ''), 'B', ''), 'C', ''), 'D', ''), 'E', '') AS INT) AS [CI_Ref],
            CONCAT(REPLACE(INDENT_NO, ']', ''), '_', REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(SECTOR_NO, 'A', ''), 'B', ''), 'C', ''), 'D', ''), 'E', '')) AS [REF],
            CAST(DIRECTION AS FLOAT) AS [Azimuth(°)],
            CAST(COORDINATE_Y AS FLOAT) AS [Latitude],
            CAST(COORDINATE_X AS FLOAT) AS [Longitude]
        FROM 
            SITE_HANDLER.dbo.radio_cell_data
        WHERE 
            Band NOT LIKE 'NBI%'
            AND LEFT(INDENT_NO, 6) IN ({}); 
        """.format(', '.join(f"'{site}'" for site in site_list))

        # Step 5: Read SQL credentials and connect
        credentials = read_credentials_from_excel_Sitedatabse(input_user)
        conn = pymssql.connect(server=credentials['server'], user=credentials['user'], password=credentials['password'])

        # Suppress warnings during query execution
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            # Step 6: Execute SQL query
            query_df = execute_query2(conn, sql_query2)

            # Step 7: Write the results to the Excel template
            with pd.ExcelWriter(template_file, engine='openpyxl', mode='a') as writer:
                query_df.to_excel(writer, sheet_name='Site DataBase', index=False)

        print("Filtered data has been saved to 'Site DataBase' sheet in the template file.")

    except pymssql.OperationalError as e:
        print(f"Database connection error: {e}")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")

def execute_query3(connection, query3):
    """Executes the given SQL query and returns the result as a DataFrame."""
    return pd.read_sql(query3, connection)

def Installation_Azimuth(Site_ID):
    try:
        sql_query3 = """
        SELECT 
            CONCAT(site, '_', sector) AS Source,
            antenna_azimuth_pre AS Pre_Planning,
            antenna_azimuth_post AS Post_Planning
        FROM 
            symon_db.dbo.nemo_installation_changes
        WHERE 
            site = %s; 
        """

        credentials = read_credentials_from_excel_Sitedatabse(input_user)
        conn = pymssql.connect(server=credentials['server'], user=credentials['user'], password=credentials['password'])
        
        # Suppress warnings during query execution
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            
            # Create a cursor and execute the query with parameters
            with conn.cursor() as cursor:
                cursor.execute(sql_query3, (Site_ID,))  # Pass the parameter as a tuple
                columns = [column[0] for column in cursor.description]  # Get column names
                query_df = pd.DataFrame(cursor.fetchall(), columns=columns)  # Fetch all results into a DataFrame

            # Write results to the Excel template
            with pd.ExcelWriter(template_file, engine='openpyxl', mode='a') as writer:
                if 'Azimuth' in writer.book.sheetnames:
                    std = writer.book['Azimuth']
                    writer.book.remove(std)
                query_df.to_excel(writer, sheet_name='Azimuth', index=False)

        print("Filtered data has been saved to 'Azimuth' sheet in the template file.")

    except pymssql.OperationalError as e:
        print(f"Database connection error: {e}")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
    finally:
        conn.close()  # Ensure the connection is closed

        
"""Calculate the bearing angle between two geographical points."""        
def calculate_bearing(lat1, lon1, lat2, lon2):

    try:
        lat1_rad = math.radians(lat1)
        lon1_rad = math.radians(lon1)
        lat2_rad = math.radians(lat2)
        lon2_rad = math.radians(lon2)
        
        delta_lon = lon2_rad - lon1_rad
        
        x = math.sin(delta_lon) * math.cos(lat2_rad)
        y = (math.cos(lat1_rad) * math.sin(lat2_rad) - 
             math.sin(lat1_rad) * math.cos(lat2_rad) * math.cos(delta_lon))
        
        bearing = math.atan2(x, y)
        bearing_deg = math.degrees(bearing)
        bearing_deg = (bearing_deg + 360) % 360
        
        return bearing_deg
    except Exception as e:
        print(f"Error calculating bearing: {e}")
        return np.nan

"""Perform vlookup in a DataFrame."""
def vlookup(values, lookup_df, lookup_col, result_col):
    lookup_df.columns = lookup_df.columns.str.strip()
    if lookup_col not in lookup_df.columns or result_col not in lookup_df.columns:
        raise KeyError(f"Columns {lookup_col} or {result_col} are missing from the lookup DataFrame.")
    
    lookup_dict = lookup_df.set_index(lookup_col)[result_col].to_dict()
    return values.map(lookup_dict)

"""Apply formulas to DataFrame based on site database."""
def apply_Formulas(df, site_db_df):
    df = df.copy()
    df['Source'] = df['Source'].astype(str)
    df['Target'] = df['Target'].astype(str)
    site_db_df.columns = site_db_df.columns.str.strip()

    required_columns = ['Site ID', 'Latitude', 'Longitude', 'REF', 'Azimuth(°)']
    missing_columns = [col for col in required_columns if col not in site_db_df.columns]
    if missing_columns:
        raise KeyError(f"Missing columns in the lookup DataFrame: {', '.join(missing_columns)}")

    # Calculate Bearing Angle
    df['Latitude1'] = vlookup(df['Source'].str[:6], site_db_df, 'Site ID', 'Latitude')
    df['Longitude1'] = vlookup(df['Source'].str[:6], site_db_df, 'Site ID', 'Longitude')
    df['Azimuth(°)1'] = vlookup(df['Source'].str[:6] + "_" + df['Source'].str[-1], site_db_df, 'REF', 'Azimuth(°)')
    
    df['Latitude2'] = vlookup(df['Target'].str[:6], site_db_df, 'Site ID', 'Latitude')
    df['Longitude2'] = vlookup(df['Target'].str[:6], site_db_df, 'Site ID', 'Longitude')
    df['Azimuth(°)2'] = vlookup(df['Target'].str[:6] + "_" + df['Target'].str[-1], site_db_df, 'REF', 'Azimuth(°)')

    # Convert to numeric
    df[['Azimuth(°)1', 'Azimuth(°)2']] = df[['Azimuth(°)1', 'Azimuth(°)2']].apply(pd.to_numeric, errors='coerce')
    df[['Latitude1', 'Longitude1', 'Latitude2', 'Longitude2']] = df[['Latitude1', 'Longitude1', 'Latitude2', 'Longitude2']].apply(pd.to_numeric, errors='coerce')

    df['Angle1'] = (df['Azimuth(°)1'] - 90) % 360
    df['Angle2'] = (df['Azimuth(°)1'] + 90) % 360

    df['Bearing angle'] = df.apply(
        lambda row: calculate_bearing(row['Latitude1'], row['Longitude1'], row['Latitude2'], row['Longitude2'])
        if pd.notna(row['Latitude1']) and pd.notna(row['Latitude2']) else np.nan, axis=1
    )

    # Fix the calculation of Delta(A1-Bearing Angle)
    df['Delta(A1-Bearing Angle)'] = df.apply(
        lambda row: np.nan if pd.isna(row['Bearing angle']) else
        (row['Bearing angle'] - row['Angle1']) % 360, axis=1
    )

    df['Delta(A1-A2)'] = abs(df['Angle1'] - df['Angle2'])
    df['Delta(A1-A2)'] = df['Delta(A1-A2)'].apply(lambda x: x if x <= 180 else 360 - x)

    # Update Co-Sited column based on conditions and include Right[Target,1]
    df['Result'] = np.where(
        (df['Source'].str[:6] == df['Target'].str[:6]) & (df['Source'].str[:9] == df['Target'].str[:9]),
        'a_Intra S' + df['Target'].str[-1],  # Append last character of Target
        np.where(
            (df['Source'].str[:6] == df['Target'].str[:6]),
            'b_Inter S' + df['Target'].str[-1],  # Append last character of Target
            np.where(
                df['Delta(A1-Bearing Angle)'] < df['Delta(A1-A2)'],  # Use the column reference directly
                'INBEAM',
                'OUTOFBEAM'
            )
        )
    )

    # Add the Quadrant column with updated logic
    df['Quadrant'] = np.where(
        df['Source'].str[:6] == df['Target'].str[:6],  # Check if Source and Target are in the same sector
        'Invalid Quadrant',
        df['Bearing angle'].apply(
            lambda x: 'Quadrant I' if 0 <= x < 90 else 
            'Quadrant II' if 90 <= x < 180 else 
            'Quadrant III' if 180 <= x < 270 else 
            'Quadrant IV' if 270 <= x < 360 else 'Invalid Quadrant'
        )
    )

    return df

"""Create LTE_HO_Analysis.xlsx based on existing data."""
def lte_ho_analysis_creation():
    
    # Step 1: Copy 'Pivot.xlsx' to 'LTE_HO_Analysis.xlsx', rename 'Sheet1' to 'Summary1', and add 'Direction' and 'Quadrant1' columns
    pivot_df = pd.read_excel(pivot_output, sheet_name='Sheet1')
    pivot_df['Direction'] = None  # Adding new column 'Direction'
    pivot_df['Quadrant1'] = None  # Adding new column 'Quadrant1'

    pivot_df.to_excel(lte_ho_analysis_file, sheet_name='Summary1', index=False)

    # Step 2: Concatenate 'Source' and 'Target' as 'REF' in 'LTE_HO_Analysis.xlsx'
    df_summary = pd.read_excel(lte_ho_analysis_file, sheet_name='Summary1')
    df_summary['REF'] = df_summary['Source'].astype(str) + '-' + df_summary['Target'].astype(str)

    # Step 3: Concatenate 'Source' and 'Target' as 'REF' in 'Direction.xlsx'
    df_pre_mod = pd.read_excel(direction_file, sheet_name='Pre-Modernization')
    df_pre_mod['REF'] = df_pre_mod['Source'].astype(str) + '-' + df_pre_mod['Target'].astype(str)

    # Step 4: Merge dataframes and update 'Direction' and 'Quadrant1' columns
    merged_df = df_summary.merge(df_pre_mod[['REF', 'Result', 'Quadrant']], on='REF', how='left')
    merged_df['Direction'] = merged_df['Result']
    merged_df['Quadrant1'] = merged_df['Quadrant']

    # Drop the 'Result' and 'Quadrant' columns as they are no longer needed
    merged_df.drop(columns=['REF', 'Result', 'Quadrant1'], inplace=True)

    # Save the updated DataFrame to 'LTE_HO_Analysis.xlsx'
    with pd.ExcelWriter(lte_ho_analysis_file, engine='openpyxl', mode='w') as writer:
        merged_df.to_excel(writer, sheet_name='Summary1', index=False)

    print("lte_ho_analysis_file > Summary1 Sheet Prepared..")
    
"""Create a summary sheet with results in LTE_HO_Analysis Excel."""   
def lte_ho_analysis_summary_Sheet():

    df_summary1 = pd.read_excel(lte_ho_analysis_file, sheet_name='Summary1')

    # Extract the required columns
    columns_to_copy = ['Source', 'Target', 'Pre_avg_Attempts', 'Post_avg_Attempts', 'Direction']
    df_summary_copy = df_summary1[columns_to_copy]

    # Step 2: Save the extracted data to a new sheet 'Summary'
    with pd.ExcelWriter(lte_ho_analysis_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df_summary_copy.to_excel(writer, sheet_name='Summary', index=False)

    print("lte_ho_analysis_file > Summery Sheet Prepared..")

    
"""Create a x-Direction sheet with results in LTE_HO_Analysis Excel.""" 
def lte_ho_analysis_x_direction():

    # Define file paths and headers
    HEADERS = ['Source', 'Period', 'Result', 'HO Attempts', 'Quadrant']

    # Load workbook and ensure the sheet exists and headers are set
    book = load_workbook(lte_ho_analysis_file)
    if 'x_Direction' not in book.sheetnames:
        sheet = book.create_sheet('x_Direction')
        sheet.append(HEADERS)
    else:
        sheet = book['x_Direction']
        if sheet.max_row == 1 and not any(sheet.cell(row=1, column=col).value for col in range(1, len(HEADERS) + 1)):
            sheet.append(HEADERS)

    # Read and process data
    df = pd.read_excel(lte_ho_analysis_file, sheet_name='Summary1').rename(columns=lambda x: x.strip())
    pre_df = df[['Source', 'Pre_avg_Attempts', 'Direction', 'Quadrant']].assign(Period='Pre').rename(columns={'Pre_avg_Attempts': 'HO Attempts', 'Direction': 'Result'})
    post_df = df[['Source', 'Post_avg_Attempts', 'Direction', 'Quadrant']].assign(Period='Post').rename(columns={'Post_avg_Attempts': 'HO Attempts', 'Direction': 'Result'})
    combined_df = pd.concat([pre_df, post_df], ignore_index=True)[HEADERS]

    # Append data to the sheet
    for row in combined_df.itertuples(index=False, name=None):
        sheet.append(row)

    # Save the workbook
    book.save(lte_ho_analysis_file)

    print("lte_ho_analysis_file > x_Direction Sheet Prepared..")

"""Create a traffic, ho attempts, mimo ranking and pathloss trend in LTE_HO_Analysis Excel.""" 
def lte_ho_analysis_pm_data():

    # Read the data from 'MIMO.xlsx'
    df = pd.read_excel(pm_data_file, sheet_name='1')

    # Process the data
    df['Date'] = pd.to_datetime(df['time']).dt.strftime('%Y%m%d')
    df['Layer'] = df['object'].str[-12:].str[:2]
    df['Sector'] = df['object'].str[-10:].str[0].astype(int)
    df.rename(columns={'HO_Exe_Att': 'HO_Exe_Att', 'LTE_M_Traffic_Data_Volume_DL_MB': 'LTE_M_Traffic_Data_Volume_DL_MB', 'LTE_Rank2_Ratio': 'LTE_Rank2_Ratio', 'LTE_UL_Path_Loss_dB': 'LTE_UL_Path_Loss_dB'}, inplace=True)
    result_df = df[['Date', 'Layer', 'Sector', 'HO_Exe_Att', 'LTE_M_Traffic_Data_Volume_DL_MB', 'LTE_Rank2_Ratio', 'LTE_UL_Path_Loss_dB']]

    # Sort the result by date (ascending order by default)
    result_df = result_df.sort_values(by='Date')

    # Write the result into 'LTE_HO_ANALYSIS.xlsx'
    with pd.ExcelWriter(lte_ho_analysis_file, engine='openpyxl', mode='a') as writer:
        result_df.to_excel(writer, index=False, sheet_name='PM_Data', startrow=0)

    print("lte_ho_analysis_file > PM_Data Sheet Prepared..")
    
"""Create a new neighbors with % ho attempys layer and sector wise in LTE_HO_Analysis Excel.""" 
def lte_ho_analysis_new_neighbor_Post():

    # Step 1: Read the Excel file
    df = pd.read_excel(siteid_file, sheet_name='Sheet1')

    # Add Ref column
    df['Ref'] = df['Source'] + '#' + df['Target']

    # Separate data into Pre and Post periods
    pre_df = df[df['Period'] == 'Pre'].copy()  # Use .copy() to avoid SettingWithCopyWarning
    post_df = df[df['Period'] == 'Post'].copy()  # Use .copy() to avoid SettingWithCopyWarning

    # Create Ref columns
    pre_df['Ref'] = pre_df['Source'] + '#' + pre_df['Target']
    post_df['Ref'] = post_df['Source'] + '#' + post_df['Target']

    # Identify new neighbors
    new_nbr = post_df[~post_df['Ref'].isin(pre_df['Ref'])]

    # Calculate total attempts for each Source in Post period
    total_attempts_post = post_df.groupby('Source')['Attempts'].sum().reset_index(name='Total Attempts')

    # Sum attempts for new neighbors
    new_nbr_sum = new_nbr.groupby('Source')['Attempts'].sum().reset_index(name='Sum of Attempts')

    # Merge to get the total attempts for each Source
    new_nbr_summary = pd.merge(new_nbr_sum, total_attempts_post, on='Source')

    # Calculate %HO Attempts
    new_nbr_summary['%HO Attempts'] = 100 * new_nbr_summary['Sum of Attempts'] / new_nbr_summary['Total Attempts']

    # Extract Layer and Sector
    def extract_layer_sector(source):
        layer = source[-3:-1]  # Extract 2 characters from the right, skipping the last character
        sector = source[-1]  # Extract the last character
        return layer, sector

    # Apply function to create Layer and Sector columns
    new_nbr_summary[['Layer', 'Sector']] = new_nbr_summary['Source'].apply(lambda src: pd.Series(extract_layer_sector(src)))

    # Final Output
    final_output = new_nbr_summary[['Layer', 'Sector', '%HO Attempts']]

    # Step 6: Write to Excel using openpyxl
    # Load the existing workbook
    book = load_workbook(lte_ho_analysis_file)

    # Remove the existing sheet if it exists
    if 'New_Nbr' in book.sheetnames:
        del book['New_Nbr']

    # Create a new sheet and add data
    new_sheet = book.create_sheet('New_Nbr')

    # Append header
    headers = final_output.columns.tolist()
    new_sheet.append(headers)

    # Append rows
    for row in dataframe_to_rows(final_output, index=False, header=False):
        new_sheet.append(row)

    # Save the workbook
    book.save(lte_ho_analysis_file)

    print("lte_ho_analysis_file.xlsx > 'New_Nbr' Sheet Prepared..")


    # Remove the "Summary1" sheet if it exists
    if 'Summary1' in book.sheetnames:
        del book['Summary1']

    # Save the workbook
    book.save(lte_ho_analysis_file)

    print("lte_ho_analysis_file > Summery1 Sheet Deleted..")


"""Calculate the Compute_Azimuth based on pre and post HO attempts"""
def process_angle_data(input_path, output_path, is_post_modernization=False):
    # Load relevant data
    df_summary = pd.read_excel(os.path.join(input_path, 'LTE_HO_Analysis.xlsx'), sheet_name='Summary')
    mod_df = pd.read_excel(os.path.join(input_path, "Direction.xlsx"), 
                            sheet_name='Post-Modernization' if is_post_modernization else 'Pre-Modernization')

    # Filter summary data
    filtered_df = df_summary[df_summary['Direction'].isin(['INBEAM', 'OUTOFBEAM'])]

    # Set attempts column based on modernization status
    attempts_col = 'Post_avg_Attempts' if is_post_modernization else 'Pre_avg_Attempts'

    # Calculate total attempts per source
    source_sum = filtered_df.groupby('Source')[attempts_col].sum().reset_index().rename(columns={attempts_col: 'source_sum'})

    # Prepare DataFrame for azimuth calculations
    compute_azimuth_df = filtered_df[['Source', 'Target', attempts_col]].merge(source_sum, on='Source')

    # Calculate handover percentage
    compute_azimuth_df[f'%HO_{"Post" if is_post_modernization else "Pre"}'] = 100 * compute_azimuth_df[attempts_col] / compute_azimuth_df['source_sum']

    # Sort values for CDF calculation
    compute_azimuth_df.sort_values(by=['Source', f'%HO_{"Post" if is_post_modernization else "Pre"}'], ascending=[True, False], inplace=True)

    # Calculate cumulative distribution function (CDF)
    compute_azimuth_df[f'%CDF_{"Post" if is_post_modernization else "Pre"}'] = compute_azimuth_df.groupby('Source')[f'%HO_{"Post" if is_post_modernization else "Pre"}'].cumsum()

    # Filter CDF to include only the top 75%
    filtered_df_cdf = compute_azimuth_df[compute_azimuth_df[f'%CDF_{"Post" if is_post_modernization else "Pre"}'] <= 100].copy()

    # Create a reference column for merging
    filtered_df_cdf['Reference'] = filtered_df_cdf['Source'] + "_" + filtered_df_cdf['Target']
    mod_df['Reference'] = mod_df['Source'] + "_" + mod_df['Target']

    # Merge the filtered DataFrame with modernization data to get bearing angles
    filtered_df_cdf = filtered_df_cdf.merge(mod_df[['Reference', 'Bearing angle']], on='Reference', how='left')

    # Calculate weighted sine and cosine values
    filtered_df_cdf['Bearing radians'] = np.radians(filtered_df_cdf['Bearing angle'])
    filtered_df_cdf['Weighted Sine'] = np.sin(filtered_df_cdf['Bearing radians']) * filtered_df_cdf[attempts_col]
    filtered_df_cdf['Weighted Cosine'] = np.cos(filtered_df_cdf['Bearing radians']) * filtered_df_cdf[attempts_col]

    # Calculate the resultant azimuth
    grouped = filtered_df_cdf.groupby('Source').agg({'Weighted Sine': 'sum', 'Weighted Cosine': 'sum'}).reset_index()
    resultant_angles_rad = np.arctan2(grouped['Weighted Sine'], grouped['Weighted Cosine'])
    azimuth_col = 'Post_Compute_Azimuth' if is_post_modernization else 'Pre_Compute_Azimuth'
    grouped[azimuth_col] = (np.degrees(resultant_angles_rad) % 360).astype(int)

    # Merging back results
    filtered_df_cdf = filtered_df_cdf.merge(grouped[['Source', azimuth_col]], on='Source', how='left')

    # Save to Excel
    final_output_file = os.path.join(output_path, 'Compute_Azimuth.xlsx')

    # Write to Excel based on modernization status
    if is_post_modernization:
        with pd.ExcelWriter(final_output_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            filtered_df_cdf.to_excel(writer, sheet_name='Post', index=False)

            # Update the summary sheet
            try:
                pre_summary_df = pd.read_excel(final_output_file, sheet_name='Compute_Azimuth')

                # Merge and process
                pre_summary_df = pre_summary_df[['Source', 'Pre_Compute_Azimuth']].drop_duplicates()

                # Ensure Pre_Compute_Azimuth is treated as strings for splitting
                pre_summary_df['Pre_Compute_Azimuth'] = pre_summary_df['Pre_Compute_Azimuth'].astype(str).replace('nan', '')

                # Split Pre_Compute_Azimuth by '#' into separate columns
                pre_summary_split = pre_summary_df['Pre_Compute_Azimuth'].str.split('#', expand=True)
                pre_summary_split.columns = [f'Pre_Compute_Azimuth_{i+1}' for i in range(pre_summary_split.shape[1])]

                # Combine Source with the split columns
                pre_summary_combined = pd.concat([pre_summary_df[['Source']], pre_summary_split], axis=1)

                # Convert Pre_Compute_Azimuth columns to integers
                for col in pre_summary_split.columns:
                    pre_summary_combined[col] = pd.to_numeric(pre_summary_combined[col], errors='coerce').round(0).astype(pd.Int64Dtype())

                # Lookup Post_Compute_Azimuth from the 'Post' sheet
                post_summary_df = filtered_df_cdf[['Source', 'Post_Compute_Azimuth']].drop_duplicates()
                post_summary_df['Post_Compute_Azimuth'] = post_summary_df['Post_Compute_Azimuth'].astype(pd.Int64Dtype())
                summary_df = pd.merge(pre_summary_combined, post_summary_df, on='Source', how='outer')

                # Save the merged summary to 'Compute_Azimuth' sheet
                summary_df.to_excel(writer, sheet_name='Compute_Azimuth', index=False)
            except ValueError:
                post_summary_df = filtered_df_cdf[['Source', 'Post_Compute_Azimuth']].drop_duplicates().reset_index(drop=True)
                post_summary_df['Post_Compute_Azimuth'] = post_summary_df['Post_Compute_Azimuth'].astype(pd.Int64Dtype())
                post_summary_df.to_excel(writer, sheet_name='Compute_Azimuth', index=False)
    else:
        with pd.ExcelWriter(final_output_file, engine='openpyxl', mode='w') as writer:
            filtered_df_cdf.to_excel(writer, sheet_name='Pre', index=False)

            # Create the summary DataFrame for pre-computation
            summary_df = filtered_df_cdf[['Source', 'Pre_Compute_Azimuth']].drop_duplicates().reset_index(drop=True)
            summary_df['Pre_Compute_Azimuth'] = summary_df['Pre_Compute_Azimuth'].astype(pd.Int64Dtype())
            summary_df.to_excel(writer, sheet_name='Compute_Azimuth', index=False)

def add_planning_azimuth(input_path, output_path):
    """
    Step to add Pre_Planning and Post_Planning to the Summary sheet of Compute_Azimuth.xlsx.
    
    Parameters:
    - input_path: Path to the input directory containing Excel files.
    - output_path: Path to the directory where the output file will be saved.
    """
    # Read the 'Azimuth' sheet from the template file
    template_file = os.path.join(input_path, "Input_Swap_Sector_Template_4G.xlsx")
    azimuth_df = pd.read_excel(template_file, sheet_name='Azimuth')

    # Read the summary sheet from Compute_Azimuth.xlsx
    summary_df = pd.read_excel(os.path.join(output_path, 'Compute_Azimuth.xlsx'), sheet_name='Compute_Azimuth')

    # Merge to add 'Pre_Planning' and 'Post_Planning' based on 'Source'
    summary_df = summary_df.merge(
        azimuth_df[['Source', 'Pre_Planning', 'Post_Planning']], 
        on='Source', 
        how='left'
    )

    # Save the updated DataFrame back to the Excel file
    with pd.ExcelWriter(os.path.join(output_path, 'Compute_Azimuth.xlsx'), engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        summary_df.to_excel(writer, sheet_name='Compute_Azimuth', index=False)



def move_compute_azimuth_to_summary(output_path):
    """
    Move the 'Compute_Azimuth' sheet from 'Compute_Azimuth.xlsx' to 'LTE_HO_Analysis.xlsx'.
    
    Parameters:
    - output_path: Path to the directory where the output file will be saved.
    """
    compute_azimuth_path = os.path.join(output_path, 'Compute_Azimuth.xlsx')
    lte_ho_analysis_path = os.path.join(output_path, 'LTE_HO_Analysis.xlsx')

    # Load both workbooks
    compute_wb = load_workbook(compute_azimuth_path)
    compute_sheet = compute_wb['Compute_Azimuth']
    lte_wb = load_workbook(lte_ho_analysis_path)

    # Create a new sheet in the LTE_HO_Analysis workbook
    new_sheet = lte_wb.create_sheet('Compute_Azimuth')

    # Append data from the compute sheet to the new sheet
    for row in compute_sheet.iter_rows(values_only=True):
        new_sheet.append(row)

    # Save the updated LTE_HO_Analysis workbook
    lte_wb.save(lte_ho_analysis_path)

"""Rename output file with siteid """  
def rename_excel_file(site_id):

    if site_id is None:
        print("Site_ID is not defined. Exiting.")
        return

    # Define the source and destination file paths
    source_file = os.path.join(input_path, 'Input_LTE_Swap_Dashboard.xlsm')
    new_filename = f"{site_id}.xlsm"
    destination_file = os.path.join(input_path, new_filename)
    
    try:
        # Ensure the source file exists
        if not os.path.exists(source_file):
            raise FileNotFoundError(f"Source file 'Input_LTE_Swap_Dashboard.xlsm' not found!")
        
        # Copy and rename the file
        shutil.copy(source_file, destination_file)
        print(f"final output ready : {new_filename}")
    
    except Exception as e:
        print(f"An error occurred: {e}")
        
"""Delete specific unwanted files if they exist."""        
def delete_unwanted_Files():

    files_to_delete = [
        "Source_CI.xlsx",
        "Pre_Post.xlsx",
        "SiteID.xlsx",
        "Pivot.xlsx",
        "Input_Swap_Sector_Template_4G.xlsx",
        "Direction.xlsx" ,
        "Compute_Azimuth.xlsx"
        
    ]

    """Check if a file is locked by another process."""
    def is_file_locked(file_path):
        return os.path.exists(file_path) and os.path.isfile(file_path) and os.access(file_path, os.W_OK) == False

    """Terminate all Excel processes to free locked files."""
    def terminate_excel_processes(file_path):
        for proc in psutil.process_iter(['pid', 'name', 'cmdline']):
            try:
                if 'EXCEL.EXE' in proc.name():
                    cmdline = proc.cmdline()
                    if cmdline and file_path in ' '.join(cmdline):
                        proc.terminate()
            except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
                pass

    """Try again to delete the unwanted files."""
    def delete_file_with_retry(file_path, max_retries=5, retry_interval=1):
        retries = 0
        while retries < max_retries:
            try:
                os.remove(file_path)
                print(f"Successfully deleted {file_path}")
                break
            except FileNotFoundError:
                print(f"{file_path} not found")
                break
            except PermissionError:
                print(f"{file_path} is currently in use, waiting to retry...")
                time.sleep(retry_interval)
                retries += 1

    """Loop through each file for deletion"""
    for file_to_delete in files_to_delete:
        file_path = os.path.join(input_path, file_to_delete)
        if is_file_locked(file_path):
            print(f"{file_to_delete} is currently in use and cannot be deleted. Closing Excel instances holding the file...")
            terminate_excel_processes(file_path)
            print(f"Excel instances holding {file_to_delete} closed. Retrying to delete {file_to_delete}...")

        """Attempt to delete the file with retry mechanism"""
        delete_file_with_retry(file_path)

"""Main function to execute all data processing steps in sequence."""
if __name__ == "__main__":

    try:
        """Basic steps to process input files for analysis."""
        process_data()
        Site_ID = get_site_id()
        define_pre_post_periods()
        delete_buffer_periods_from_input()
        Deltion_Zero_Attempts()
        eNodeBID_Fetch()
        neighbour_gcid_siteid()
        missing_gcid_enode_database()
        prepare_ho_pre_post_relationwise()
        site_database()

        """Bearing Angle calculation based on lat-Long and azimuth, Inter and Intra HO%."""
        pivot_df = pd.read_excel(pivot_output, sheet_name='Sheet1')
        site_db_df = pd.read_excel(template_file, sheet_name='Site DataBase')
        Installation_Azimuth(Site_ID)
        Installation_Azimuth_df = pd.read_excel(template_file, sheet_name='Azimuth')

        df1 = pivot_df[['Source', 'Target', 'Pre_avg_Attempts']]
        df2 = pivot_df[['Source', 'Target', 'Post_avg_Attempts']]

        """Apply fomulas to get data from Input_Swap_Sector_Template_4G to calculate bearing angle and inter & intra."""
        df1 = apply_Formulas(df1, site_db_df)
        df2 = apply_Formulas(df2, site_db_df)

        """Write final output data in Direction.xlsx file."""
        with pd.ExcelWriter(direction_file, engine='openpyxl') as writer:
            df1.to_excel(writer, sheet_name='Pre-Modernization', index=False)
            df2.to_excel(writer, sheet_name='Post-Modernization', index=False)

        print("direction.xlsx has been successfully created.")

        """Last file LTE_HO_Analysis sheet preperation."""
        lte_ho_analysis_creation()
        lte_ho_analysis_summary_Sheet()
        lte_ho_analysis_x_direction()
        lte_ho_analysis_pm_data()
        lte_ho_analysis_new_neighbor_Post()
        
        """Computing Pre and Post azimuth based on Ho attempts direction"""
        process_angle_data(input_path, output_path, is_post_modernization=False)  # Pre-computation
        process_angle_data(input_path, output_path, is_post_modernization=True)   # Post-computation
        add_planning_azimuth(input_path, output_path)  # Add Planning_Azimuth
        move_compute_azimuth_to_summary(output_path)  # Move Compute_Azimuth to LTE_HO_Analysis
        print("Compute_Azimuth files prepared and moved to LTE_HO_Analysis...")

        """Renaming output file with siteid."""
        
        print(f"site_id: {Site_ID}")
        rename_excel_file(Site_ID)

        """cleanup the folder delete unwanted file."""
        delete_unwanted_Files()

    except Exception as e:
        print(f"An error occurred: {e}")
    
